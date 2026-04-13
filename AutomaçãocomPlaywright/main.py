from tkinter import Tk, Label, Text, Scrollbar, Button, Frame, Entry, font as tkfont
from tkinter.filedialog import askopenfilename
from playwright.sync_api import sync_playwright
import openpyxl
import time
import threading
import random

class Interface:
    def __init__(self):
        self.janela = Tk()
        self.janela.title("Automacao - Extrair Titulos")
        self.janela.geometry("900x650")
        self.janela.configure(bg="#1a1a2e")
        self.janela.resizable(False, False)

        self.style_title = {"font": ("Helvetica", 18, "bold"), "bg": "#16213e", "fg": "#e94560", "padx": 20, "pady": 10}
        self.style_label = {"font": ("Helvetica", 10), "bg": "#16213e", "fg": "#eaeaea"}
        self.style_button = {"font": ("Helvetica", 10, "bold"), "cursor": "hand2", "relief": "flat", "bd": 0}
        self.style_text = {"font": ("Courier", 9), "bg": "#0f0f23", "fg": "#00ff88", "insertbackground": "#00ff88"}

        self.frame_titulo = Frame(self.janela, bg="#16213e")
        self.frame_titulo.pack(fill="x")
        self.label_titulo = Label(self.frame_titulo, text="Automacao - Extrair Titulos", **self.style_title)
        self.label_titulo.pack(pady=15)

        self.frame_botoes = Frame(self.janela, bg="#16213e")
        self.frame_botoes.pack(pady=15)

        self.btn_importar = Button(self.frame_botoes, text="📁 Importar", width=14, bg="#4361ee", fg="white", command=self.importar, **self.style_button)
        self.btn_importar.pack(side="left", padx=4)

        self.btn_iniciar = Button(self.frame_botoes, text="▶ Iniciar", width=10, bg="#2ec4b6", fg="white", state="disabled", command=self.iniciar, **self.style_button)
        self.btn_iniciar.pack(side="left", padx=4)

        self.btn_pausar = Button(self.frame_botoes, text="⏸ Pausar", width=9, bg="#f77f00", fg="white", state="disabled", command=self.pausar, **self.style_button)
        self.btn_pausar.pack(side="left", padx=4)

        self.btn_encerrar = Button(self.frame_botoes, text="⏹ Encerrar", width=9, bg="#d62828", fg="white", state="disabled", command=self.encerrar, **self.style_button)
        self.btn_encerrar.pack(side="left", padx=4)

        self.label_status = Label(self.janela, text="Aguardando importacao...", font=("Helvetica", 12, "bold"), bg="#1a1a2e", fg="#f8f8f2")
        self.label_status.pack(pady=5)

        self.frame_timer = Frame(self.janela, bg="#16213e", relief="sunken", bd=2)
        self.frame_timer.pack(pady=5, padx=20, fill="x")

        self.label_tempo = Label(self.frame_timer, text="", font=("Courier", 10, "bold"), bg="#16213e", fg="#00d9ff", justify="center")
        self.label_tempo.pack(pady=6, fill="both", expand=True)

        self.frame_texto = Frame(self.janela, bg="#0f0f23", relief="sunken", bd=2)
        self.frame_texto.pack(padx=20, pady=(5, 10), fill="both", expand=True)

        self.texto = Text(self.frame_texto, width=100, height=22, **self.style_text, relief="flat", bd=0)
        self.texto.pack(side="left", fill="both", expand=True, padx=5, pady=5)

        self.scrollbar = Scrollbar(self.frame_texto, command=self.texto.yview, bg="#0f0f23")
        self.scrollbar.pack(side="right", fill="y")
        self.texto.config(yscrollcommand=self.scrollbar.set)

        self.pausado = False
        self.encerrado = False
        self.pagina = None
        self.inicio_cpf = 0
        self.tempo_total = 0
        self.processados = 0
        self.cpfs_restantes = 0
        self.caminho_arquivo = None

    def atualizar(self, mensagem):
        cor = self.get_cor_mensagem(mensagem)
        self.texto.insert("end", mensagem + "\n", cor)
        self.texto.tag_config("verde", foreground="#00ff88")
        self.texto.tag_config("amarelo", foreground="#ffd60a")
        self.texto.tag_config("vermelho", foreground="#ff6b6b")
        self.texto.tag_config("azul", foreground="#00d9ff")
        self.texto.tag_config("branco", foreground="#f8f8f2")
        self.texto.see("end")
        self.janela.update()

    def get_cor_mensagem(self, msg):
        if "SUCESSO" in msg or "sucesso" in msg.lower():
            return "verde"
        elif "ERRO" in msg or "erro" in msg.lower():
            return "vermelho"
        elif "PAUSADO" in msg or "CONTINUANDO" in msg or "ENCERRANDO" in msg:
            return "amarelo"
        elif "CPF" in msg or "Processando" in msg or "Navegando" in msg or "Selecionando" in msg:
            return "azul"
        else:
            return "branco"

    def status(self, texto):
        self.label_status.config(text=texto)
        self.janela.update()

    def atualizar_tempo(self, restantes, tempo_medio, tempo_atual=0):
        texto = f"Restantes: {restantes} | Media: {tempo_medio:.1f}s | Atual: {tempo_atual:.1f}s"
        if tempo_medio > 0 and restantes > 0:
            segundos_restantes = restantes * tempo_medio
            horas = int(segundos_restantes // 3600)
            minutos = int((segundos_restantes % 3600) // 60)
            segundos = int(segundos_restantes % 60)
            if horas > 0:
                texto += f" | Resto: ~{horas}h {minutos}m"
            elif minutos > 0:
                texto += f" | Resto: ~{minutos}m {segundos}s"
            else:
                texto += f" | Resto: ~{segundos}s"
        self.label_tempo.config(text=texto)
        self.janela.update()

    def atualizar_tempo_realtime(self):
        while not self.encerrado:
            if self.inicio_cpf > 0:
                tempo_atual = time.time() - self.inicio_cpf
                tempo_medio = self.tempo_total / self.processados if self.processados > 0 else 0
                self.atualizar_tempo(self.cpfs_restantes, tempo_medio, tempo_atual)
            time.sleep(0.1)

    def esperar(self, segundos=1, aleatorio=False):
        if aleatorio:
            segundos += random.uniform(0.1, 0.3)
        inicio = time.time()
        while time.time() - inicio < segundos:
            tempo_restante = segundos - (time.time() - inicio)
            if tempo_restante > 0.05:
                time.sleep(min(0.05, tempo_restante))
            self.janela.update()
            if self.encerrado:
                break

    def importar(self):
        caminho_arquivo = askopenfilename(
            title="Selecione a planilha Excel",
            filetypes=[("Arquivos Excel", "*.xlsx")]
        )
        if caminho_arquivo:
            self.caminho_arquivo = caminho_arquivo
            self.atualizar(f"Arquivo selecionado: {caminho_arquivo}")
            self.btn_importar.config(state="disabled", bg="#555")
            self.btn_iniciar.config(state="normal", bg="#2ec4b6")
            self.status("Planilha importada - Clique em Iniciar")

    def iniciar(self):
        if not self.caminho_arquivo:
            self.atualizar("Importe uma planilha primeiro!")
            return
        self.btn_iniciar.config(state="disabled", bg="#555")
        self.btn_pausar.config(state="normal", bg="#f77f00")
        self.btn_encerrar.config(state="normal", bg="#d62828")
        threading.Thread(target=self.executar, args=(self.caminho_arquivo,), daemon=True).start()

    def pausar(self):
        if not self.pausado:
            self.pausado = True
            self.btn_pausar.config(text="▶ Continuar", bg="#2ec4b6")
            self.status("PAUSADO")
            self.atualizar(">>> PAUSADO <<<")
        else:
            self.pausado = False
            self.btn_pausar.config(text="⏸ Pausar", bg="#f77f00")
            self.status("Continuando...")
            self.atualizar(">>> CONTINUANDO <<<")

    def aguardar_despausar(self):
        while self.pausado and not self.encerrado:
            self.janela.update()
            time.sleep(0.2)

    def encerrar(self):
        self.encerrado = True
        self.pausado = False
        self.atualizar(">>> ENCERRANDO... <<<")
        if self.pagina:
            try:
                self.pagina.goto("about:blank")
            except:
                pass
        self.status("Encerrado")
        self.janela.destroy()

    def executar(self, caminho):
        self.pausado = False
        self.encerrado = False
        self.inicio_cpf = 0
        self.tempo_total = 0
        self.processados = 0

        thread_timer = threading.Thread(target=self.atualizar_tempo_realtime, daemon=True)
        thread_timer.start()

        self.status("Carregando planilha...")
        self.atualizar("Carregando planilha...")

        try:
            wb = openpyxl.load_workbook(caminho)
            planilha = wb.active
            self.atualizar(f"Planilha: {planilha.title}")
            self.atualizar(f"Total de linhas: {planilha.max_row}")
        except Exception as e:
            self.atualizar(f"ERRO ao carregar: {e}")
            self.status("Erro")
            self.resetar_botoes()
            return

        self.status("Iniciando navegador...")
        self.atualizar("Iniciando navegador...")

        with sync_playwright() as p:
            navegador = p.chromium.launch(
                headless=False,
                args=[
                    "--disable-blink-features=AutomationControlled",
                    "--no-sandbox",
                    "--disable-dev-shm-usage"
                ]
            )
            contexto = navegador.new_context(
                user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64)"
            )

            self.pagina = pagina = contexto.new_page()

            self.atualizar("Fazendo login...")
            pagina.goto("https://www.cursoemvideo.com/login/")
            pagina.wait_for_load_state("networkidle")
            pagina.fill('xpath=//*[@id="post-42350"]/div/div[1]/div/div/div[2]/div/div[1]/div/div[3]/div/div/div/form/div[3]/input', "larissaladeira42@gmail.com")
            pagina.fill('xpath=//*[@id="uabb-password-field"]', "girlhell1")
            pagina.locator('xpath=//*[@id="post-42350"]/div/div[1]/div/div/div[2]/div/div[1]/div/div[3]/div/div/div/form/div[7]/div/button').click()
            pagina.wait_for_load_state("networkidle")
            self.esperar(1.5)
            self.atualizar("Login realizado!")
            pagina.goto("https://www.cursoemvideo.com/meus-cursos/")
            self.esperar(1.5)

            total_linhas = planilha.max_row
            total_cpfs = 0
            for linha_count in range(2, total_linhas + 1):
                cpf_temp = planilha.cell(row=linha_count, column=1).value
                titulo_temp = planilha.cell(row=linha_count, column=2).value
                if cpf_temp and str(cpf_temp).strip() != "" and (not titulo_temp or str(titulo_temp).strip() == ""):
                    total_cpfs += 1

            self.atualizar(f"CPFs para processar: {total_cpfs}")

            linha_inicio = 2
            for teste_linha in range(2, min(total_linhas + 1, 20)):
                teste_cpf = planilha.cell(row=teste_linha, column=1).value
                if teste_cpf and str(teste_cpf).strip() != "":
                    linha_inicio = teste_linha
                    break

            self.cpfs_restantes = total_cpfs
            self.atualizar_tempo(self.cpfs_restantes, 0, 0)

            for linha_atual in range(linha_inicio, total_linhas + 1):
                if self.encerrado:
                    break

                self.aguardar_despausar()

                cpf = planilha.cell(row=linha_atual, column=1).value

                if not cpf or str(cpf).strip() == "":
                    continue

                cpf_str = str(cpf).strip()
                if cpf_str == "":
                    continue

                cpf_formatado = cpf_str.zfill(11)
                self.status(f"Processando linha {linha_atual}")
                self.atualizar(f"[{linha_atual}/{total_linhas}] Processando CPF: {cpf_formatado}")

                self.inicio_cpf = time.time()
                self.cpfs_restantes -= 1

                try:
                    margem_atual = planilha.cell(row=linha_atual, column=2).value

                    if margem_atual and str(margem_atual).strip() != "":
                        self.atualizar(f"[{linha_atual}] Ja possui titulo: '{margem_atual}' - PULANDO")
                        tempo_cpf = time.time() - self.inicio_cpf
                        self.tempo_total += tempo_cpf
                        self.processados += 1
                        self.inicio_cpf = 0
                        continue

                    self.atualizar(f"[{linha_atual}] Navegando para cursos...")
                    pagina.locator('xpath=//*[@id="menu-item-283541"]/a').click()
                    pagina.wait_for_load_state("networkidle")
                    self.esperar(1, aleatorio=True)

                    self.atualizar(f"[{linha_atual}] Selecionando curso...")
                    pagina.locator('xpath=//*[@id="post-283493"]/div/div[1]/div/div/div[2]/div/div/div/div[3]/div/div[1]/div[13]/div/div[5]/a/img').click()
                    pagina.wait_for_load_state("networkidle")
                    self.esperar(1, aleatorio=True)

                    self.atualizar(f"[{linha_atual}] Expandindo modulo...")
                    pagina.locator('xpath=//*[@id="ld-expand-button-26338"]/span[2]').click()
                    self.esperar(0.5, aleatorio=True)

                    self.atualizar(f"[{linha_atual}] Abrindo aula...")
                    pagina.locator('xpath=//*[@id="ld-table-list-item-26340"]/a/span').click()
                    pagina.wait_for_load_state("networkidle")
                    self.esperar(1, aleatorio=True)

                    xpath_titulo = 'xpath=//*[@id="ld-tab-content-26338"]/h2[1]'
                    pagina.wait_for_selector(xpath_titulo, timeout=5000)
                    titulo = pagina.inner_text(xpath_titulo).strip()

                    planilha.cell(row=linha_atual, column=2, value=titulo)
                    wb.save(caminho)
                    tempo_cpf = time.time() - self.inicio_cpf
                    self.tempo_total += tempo_cpf
                    self.processados += 1
                    self.inicio_cpf = 0
                    self.atualizar(f"[{linha_atual}] SUCESSO - Titulo salvo: {titulo}")

                    pagina.goto("https://www.cursoemvideo.com/meus-cursos/")
                    self.esperar(1.5)

                except Exception as e:
                    self.atualizar(f"[{linha_atual}] ERRO: {e}")
                    tempo_cpf = time.time() - self.inicio_cpf
                    self.tempo_total += tempo_cpf
                    self.processados += 1
                    self.inicio_cpf = 0
                    try:
                        pagina.goto("https://www.cursoemvideo.com/meus-cursos/")
                    except:
                        pass
                    continue

            self.atualizar("")
            self.atualizar("=" * 50)
            self.atualizar("PROCESSO FINALIZADO")
            self.atualizar(f"Total processado: {self.processados}")
            self.atualizar("=" * 50)
            self.status("Finalizado")
            self.inicio_cpf = 0
            tempo_medio_final = self.tempo_total / self.processados if self.processados > 0 else 0
            self.atualizar_tempo(0, tempo_medio_final, 0)
            time.sleep(1)

        self.resetar_botoes()

    def resetar_botoes(self):
        self.btn_importar.config(state="normal", bg="#4361ee")
        self.btn_iniciar.config(state="disabled", bg="#555")
        self.btn_pausar.config(state="disabled", text="⏸ Pausar", bg="#f77f00")
        self.btn_encerrar.config(state="disabled", bg="#d62828")
        self.caminho_arquivo = None

    def mostrar(self):
        self.janela.mainloop()


if __name__ == "__main__":
    ui = Interface()
    ui.mostrar()
